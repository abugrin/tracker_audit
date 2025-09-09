"""Export functionality for audit results to Excel format."""

import logging
from datetime import datetime
from pathlib import Path
from typing import List

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from rich.console import Console

from audit import QueueInfo, AccessInfo
from translations import t

console = Console()
logger = logging.getLogger(__name__)


class ExcelExporter:
    """Exports audit results to Excel format."""
    
    def __init__(self):
        """Initialize the exporter."""
        self.workbook = Workbook()
    
    def export_audit_results(
        self, 
        queues: List[QueueInfo], 
        access_info: List[AccessInfo],
        output_path: Path
    ) -> bool:
        """Export audit results to Excel file.
        
        Args:
            queues: List of queue information
            access_info: List of access information
            output_path: Path to save the Excel file
            
        Returns:
            True if export successful, False otherwise
        """
        try:
            console.print(t("exporting_to_excel"), style="blue")
            
            # Remove default sheet
            if "Sheet" in self.workbook.sheetnames:
                self.workbook.remove(self.workbook["Sheet"])
            
            # Create sheets
            self._create_summary_sheet(queues, access_info)
            self._create_queues_sheet(queues)
            self._create_access_sheet(access_info)
            
            # Save workbook
            self.workbook.save(output_path)
            
            console.print(t("export_success", path=output_path), style="green")
            return True
            
        except Exception as e:
            console.print(t("export_error", error=str(e)), style="red")
            logger.error(t("export_error", error=str(e)))
            return False
    
    def _create_summary_sheet(self, queues: List[QueueInfo], access_info: List[AccessInfo]) -> None:
        """Create summary sheet with overview statistics.
        
        Args:
            queues: List of queue information
            access_info: List of access information
        """
        ws = self.workbook.create_sheet("Summary", 0)
        
        # Title
        ws['A1'] = "Yandex Tracker Audit Summary"
        ws['A1'].font = Font(size=16, bold=True)
        ws['A1'].fill = PatternFill(start_color="4472C4", end_color="4472C4", patternType="solid")
        ws['A1'].font = Font(color="FFFFFF", size=16, bold=True)
        
        # Audit metadata
        row = 3
        ws[f'A{row}'] = "Audit Date:"
        ws[f'B{row}'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ws[f'A{row}'].font = Font(bold=True)
        
        row += 1
        ws[f'A{row}'] = "Total Queues:"
        ws[f'B{row}'] = len(queues)
        ws[f'A{row}'].font = Font(bold=True)
        
        row += 1
        ws[f'A{row}'] = "Total Access Entries:"
        ws[f'B{row}'] = len(access_info)
        ws[f'A{row}'].font = Font(bold=True)
        
        # Access by subject type
        row += 2
        ws[f'A{row}'] = "Access Breakdown by Subject Type:"
        ws[f'A{row}'].font = Font(bold=True, underline='single')
        
        subject_types = {}
        for access in access_info:
            subject_types[access.subject_type] = subject_types.get(access.subject_type, 0) + 1
        
        for subject_type, count in subject_types.items():
            row += 1
            ws[f'A{row}'] = f"  {subject_type.title()}:"
            ws[f'B{row}'] = count
        
        # Queue statistics
        row += 2
        ws[f'A{row}'] = "Queue Statistics:"
        ws[f'A{row}'].font = Font(bold=True, underline='single')
        
        queues_with_access = len(set(access.queue_key for access in access_info))
        queues_without_access = len(queues) - queues_with_access
        
        row += 1
        ws[f'A{row}'] = "  Queues with access entries:"
        ws[f'B{row}'] = queues_with_access
        
        row += 1
        ws[f'A{row}'] = "  Queues without access entries:"
        ws[f'B{row}'] = queues_without_access
        
        # Auto-fit columns
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_queues_sheet(self, queues: List[QueueInfo]) -> None:
        """Create queues sheet with detailed queue information.
        
        Args:
            queues: List of queue information
        """
        ws = self.workbook.create_sheet("Queues")
        
        # Headers
        headers = ["Queue Key", "Name", "Description", "Lead", "Default Type", "Default Priority"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", patternType="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Data
        for row, queue in enumerate(queues, 2):
            ws.cell(row=row, column=1, value=queue.key)
            ws.cell(row=row, column=2, value=queue.name)
            ws.cell(row=row, column=3, value=queue.description or "")
            ws.cell(row=row, column=4, value=queue.lead or "")
            ws.cell(row=row, column=5, value=queue.default_type or "")
            ws.cell(row=row, column=6, value=queue.default_priority or "")
        
        # Auto-fit columns
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 80)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Freeze header row
        ws.freeze_panes = "A2"
    
    def _create_access_sheet(self, access_info: List[AccessInfo]) -> None:
        """Create access sheet with detailed access information.
        
        Args:
            access_info: List of access information
        """
        ws = self.workbook.create_sheet("Access Permissions")
        
        # Headers
        headers = ["Queue Key", "Permission Type", "Subject Type", "Subject ID", 
                  "Subject Display", "Granted Permissions"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", patternType="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Data
        for row, access in enumerate(access_info, 2):
            ws.cell(row=row, column=1, value=access.queue_key)
            ws.cell(row=row, column=2, value=access.permission_type)
            ws.cell(row=row, column=3, value=access.subject_type)
            ws.cell(row=row, column=4, value=access.subject_id)
            ws.cell(row=row, column=5, value=access.subject_display or "")
            ws.cell(row=row, column=6, value=", ".join(access.granted_permissions))
        
        # Auto-fit columns
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 60)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Freeze header row
        ws.freeze_panes = "A2"
        
        # Add filters
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(access_info) + 1}"
